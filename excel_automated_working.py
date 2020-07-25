import openpyxl, re, time, datetime
from openpyxl.styles import *

sama = {"sam":20}
'''

#import excel workbook & work on sheet EmployeeData then the others automated
##xl_Path = "C:/Users/Sambou/Desktop/Employees.xlsx"
##work_book = openpyxl.load_workbook(xl_Path)

#sheet EmployeeData
sheet_Emp = work_book[work_book.sheetnames[0]]

#Define a regular expression that will collect the dimensions and begin auto
PATTERN = r"(\w)(\d+):(\w)(\d+)"
STRING = sheet_Emp.dimensions
dimension_re_search = re.search(PATTERN, STRING, re.S)
#TESTING
print(dimension_re_search.group(0))

#define highest point and lowest point of out sheet
HIGHEST_ROW_VALUE = dimension_re_search.group(4)
HIGHEST_COLUMN_VALUE = dimension_re_search.group(3)
LOWEST_ROW_VALUE = dimension_re_search.group(2)
LOWEST_COLUMN_VALUE = dimension_re_search.group(1)
#TESTING
print(HIGHEST_ROW_VALUE)

'''
        
#Excel class               
class Modifier:
    "Modifies a given Excel xlsx file"
    
    def __init__(self, SheetName, FILE = "C:/Users/Sambou/Documents/Employees.xlsx"):
        
        self.SheetName = SheetName    
        self.FILE = FILE
        print(SheetName)

        global worksheets
        #excelWork.saveFunc() #TEST FALIED
        #go through all the sheet and apply some settings
        worksheets = work_book[SheetName]

        #Define a regular expression that will collect the dimensions and begin auto
        PATTERN = r"(\w)(\d+):(\w)(\d+)"
        STRING = worksheets.dimensions
        dimension_re_search = re.search(PATTERN, STRING, re.S)
        
        #print(dimension_re_search.group(0)) #TESTING SUCCESS
        #define highest point and lowest point of out sheet
        HIGHEST_ROW_VALUE = dimension_re_search.group(4)
        HIGHEST_COLUMN_VALUE = dimension_re_search.group(3)
        LOWEST_ROW_VALUE = dimension_re_search.group(2)
        LOWEST_COLUMN_VALUE = dimension_re_search.group(1) 
        #print(HIGHEST_ROW_VALUE) #TESTING SUCCESS
        
        #creating a dictionary of all the cell names using the dimension
        ALPHABET = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"     #iterate over an do my pick and choose
        enum_list = list(enumerate(ALPHABET, 1))        #enumerate starting at 1 to 26
        dict_comphn_store = {key : value for key, value in enum_list}    #go throuth the enum_list list of tuples and save them in dict_comphn_store

        global storing_list
        storing_list = []
        #loop through the dictionary and using the highest row and column, return all the cells
        for key, value in dict_comphn_store.items():
            for row in range(1, int(HIGHEST_ROW_VALUE)+1): #integers 0-9
                if value <= HIGHEST_COLUMN_VALUE and key <= int(HIGHEST_ROW_VALUE):
                    #print(f"{value+str(row)}") #TEST PASSED
                    storing_list.append(f"{value+str(row)}")
                    #print("STORED") #TEST PASSED
                else:
                    break

                if key <= int(HIGHEST_ROW_VALUE):
                    pass
                else:
                    break

            
        print(storing_list) #TEST PASSED
        Modifier.Settings(self, storing_list)


    def Settings(self, cell_list):
        self.cell_list = cell_list
        
        for each_cell in cell_list:
            cell_holder = worksheets[each_cell]
            #alignment
            align = Alignment(horizontal = "left")
            cell_holder.alignment = align
            #color
            font = Font(color = colors.RED, bold = True, italic = True)
            cell_holder.font = font
            #patternfill
            fill = PatternFill(fill_type = "solid", bgColor = colors.GREEN)
            cell_holder.fill = fill

            #save settings to excel workbook
            work_book.save(self.FILE)

            
            #time.sleep(5)
            #print(f"Settings Applied to : {self.SheetName}")
            #print(f"--> {each_cell}")
            #print(cell_list)
            
            #calling the log function and writing to the txt file
            Modifier.LogFile(self, f"Settings Applied to : {self.SheetName}\n")
            Modifier.LogFile(self, f"--> {each_cell}\n")

            
            
        
    def ClearCells(self, cell_list):
        pass

    global STRING
    STRING = str(datetime.datetime.now())
    
    def LogFile(self, LogMessage):
        self.LogMessage = LogMessage

        PATTERN = r"(.{9,10}) (\d{2}:\d{2}:\d{2})(.+)$"
        
        search_date = re.search(PATTERN, STRING)
        
        date = search_date.group(1)
        current_time = search_date.group(2)
        current_time = current_time.split(":")

        PATH = f"c:/users/sambou/Documents/LogFile_{date}_{current_time[0]}.{current_time[1]}.{current_time[2]}.txt"
        with open(PATH, "a") as logfilesam:
            #logfilesam.write("______LOGS_____\n")
            logfilesam.write(LogMessage)
            #logfilesam.write("........END......")
            logfilesam.close()
        
                    



class Main():
    def __init__(self, FILE = "C:/Users/Sambou/Documents/Employees.xlsx"):
        global work_book
        self.FILE = FILE
        work_book = openpyxl.load_workbook(FILE)
        for sheets_num in range(len(work_book.sheetnames)):
                Modifier(work_book.sheetnames[sheets_num])

        #Settings() #TEST FAILED
        #Modifier.Settings(self, storing_list) #TEST FAILED
    









if __name__ == "__main__":
    modObj = Main() #excelWork()
    #modObj
    #mod = Modifier("Skills")
    #if hasattr(mod, "FILE"):
    #pass
    #else:
        
