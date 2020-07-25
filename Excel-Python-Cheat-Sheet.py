#Installing the necessary modules
python -m pip install openpyxl

#Official documentation
#https://openpyxl.readthedocs.io/en/stable/index.html



#Workbooks
#Creating a new workbook
from openpyxl import Workbook

workbook = Workbook() #Creating a new workbook object, by initializing the Workbook() class

file_path = "D:\\ExcelApp\\MyCompanyStaff.xlsx" #Setting the path to (location of) the new Excel workbook

workbook.save(file_path) #Saving the workbook

#Loading the Excel workbook .xlsx file
workbook = openpyxl.load_workbook("D:\Employees.xlsx")

#Getting the basic properties of the workbook
workbook.properties

#Getting the sheets in the workbook
workbook.worksheets

#Getting the sheet names in a workbook, as a list of strings
workbook.sheetnames

#Getting the active sheet in the workbook
workbook.active

#Referencing a sheet by its name
workbook['EmployeeData']

sheet = workbook['EmployeeData']

#Creating a new sheet in the workbook
workbook.create_sheet('TestSheet')
workbook.save("D:\Employees.xlsx") #Saving the workbook

#Removing a sheet from the workbook
sheet = workbook['TestSheet']
workbook.remove(sheet)

#or

del workbook['TestSheet']

workbook.save("D:\Employees.xlsx") #Saving the workbook

#Closing a workbook
workbook.close()


#Sheets
#Getting the title of a sheet
sheet.title

#Setting the title of a sheet
sheet.title = 'Employees'

#Renaming the default sheet to 'Employees'
sheet = workbook['Sheet']

sheet.title = 'Employees'

#Seeing all the available sheet operations
dir(sheet)

#Getting the active cell
sheet.active_cell

#Getting the dimensions of the sheet (the array populated with data)
sheet.dimensions

#Getting basic parameters about the sheet
sheet.sheet_format

#Getting the basic properties of the sheet
sheet.sheet_properties

#Getting the sheet state (visible/hidden)
sheet.sheet_state

#Getting other general information about a sheet
sheet.sheet_view

#Getting the number of rows in the sheet
sheet.max_row

#Getting the number of columns in the sheet
sheet.max_column

#Returning the rows in a sheet as tuples
for i in sheet.values:
	print(i)

    
    
#More sheet operations 
#Getting the value of a cell in a sheet, by the cell coordinates
sheet = workbook['EmployeeData']
sheet['B10'].value

#or

sheet.cell(row = 10, column = 2).value

#Selecting a cell in the sheet
cell = sheet['B10']

#Getting the row number of the cell
cell.row

#Getting the column number of the cell
cell.column

#Getting the coordinates of a cell
cell.coordinate

#Getting the data type for a cell
'''
TYPE_STRING = 's'
    TYPE_FORMULA = 'f'
    TYPE_NUMERIC = 'n'
    TYPE_BOOL = 'b'
    TYPE_NULL = 'n'
    TYPE_INLINE = 'inlineStr'
    TYPE_ERROR = 'e'
    TYPE_FORMULA_CACHE_STRING = 'str'
'''
cell.data_type

#Getting the encoding format for a cell
cell.encoding

#Setting a value for a cell
cell = sheet['B10']
cell.value = 'Test'
workbook.save("D:\Employees.xlsx") #Saving the workbook

#Getting the parent worksheet for a cell
cell.parent



#Tables
from openpyxl.worksheet.table import Table, TableStyleInfo

#Creating a table inside the sheet
table = Table(displayName = "Table", ref = "A1:G11")

#Defining a style for the table (default style name, row/column stripes)
#Choose your table style from the default styles of openpyxl
#Just type in openpyxl.worksheet.table.TABLESTYLES in the Python interpreter
style = TableStyleInfo(name = "TableStyleMedium9", showRowStripes = True, showColumnStripes = True)

#Applying the style to the table
table.tableStyleInfo = style

#Adding the newly created table to the sheet
sheet.add_table(table)



#Cell styles
#Setting the font properties for a cell
'''All the available font-related options:
font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')

Source: https://openpyxl.readthedocs.io/en/stable/styles.html            
'''
from openpyxl.styles import *

font = Font(color = colors.RED, bold = True, italic = True)

cell.font = font

#Setting the pattern/color properties for the cell
'''
fill = PatternFill(fill_type=None,
                   start_color='FFFFFFFF',
                   end_color='FF000000',
                   bgColor = 'FF000000',
                   fgColor = 'FF000000')
                   
Source: https://openpyxl.readthedocs.io/en/stable/styles.html 
'''
from openpyxl.styles import *

fill = PatternFill(fill_type = 'solid', bgColor = 'F7FE2E')

cell.fill = fill

#Setting the border properties for the cell
'''
border = Border(left=Side(border_style=None, color='FF000000'),
                right=Side(border_style=None, color='FF000000'),
                top=Side(border_style=None,color='FF000000'),
                bottom=Side(border_style=None, color='FF000000'),
                diagonal=Side(border_style=None, color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None, color='FF000000'),
                vertical=Side(border_style=None, color='FF000000'),
                horizontal=Side(border_style=None, color='FF000000'))
                
Source: https://openpyxl.readthedocs.io/en/stable/styles.html 
'''
from openpyxl.styles import *

border = Border(left = Side(border_style = 'double', color = '322FEC'), right = Side(border_style = 'double', color = '322FEC'), top = Side(border_style = 'double', color = '322FEC'), bottom = Side(border_style = 'double', color = '322FEC'))

cell.border = border

#Setting the alignment properties for the cell
'''
alignment = Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)

Source: https://openpyxl.readthedocs.io/en/stable/styles.html 
'''
from openpyxl.styles import *

align = Alignment(horizontal = 'left')

cell.alignment = align



#Applying a default style to a table
#Creating a table inside the sheet
table = Table(displayName = "Table", ref = "A1:G11")

#Defining a style for the table (default style name, row/column stripes)
#Choose your table style from the default styles of openpyxl
#Just type in openpyxl.worksheet.table.TABLESTYLES in the Python interpreter
style = TableStyleInfo(name = "TableStyleMedium9", showRowStripes = True, showColumnStripes = True)

#Applying the style to the table
table.tableStyleInfo = style

#Adding the newly created table to the sheet
sheet.add_table(table)



#Setting the cell protection
'''
protection = Protection(locked=True,
                        hidden=False)
                        
Source: https://openpyxl.readthedocs.io/en/stable/styles.html
More information about cell protection at https://openpyxl.readthedocs.io/en/stable/protection.html
'''
from openpyxl.styles import *

protection = Protection(locked = True, hidden = False)

cell.protection = protection



#More information about openpyxl's classes, methods and attributes at https://openpyxl.readthedocs.io/en/stable/